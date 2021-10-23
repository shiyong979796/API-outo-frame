from  openpyxl import load_workbook
import pymysql
import os
from configparser import ConfigParser

import logging
import config


class  Ddt_data:
    """
    1.导入
        DDT   data driver
        form  openpyxl  import  load_workbook
    2.创建类
        初始化数据  init  file_path  &  form

            加载文件 load_workbook(file_path)
            选择表单 self.load_file[form]
    3.创建方法
        获取表单第一行数据
             创建titel列表
             for循环获取  list（表单第一行数据）[0]
             把循环到的值 添加到列表
             返回出去

        获取每行的数据 以key value的形式返回
            创建 总数据列表
            for 循环获取 list（表单）[1:]
            创建行列表
            把循环到的行值添加到行列表内
            用title 做key 循环到的value 做值 all_data.append(dict(zip(self.get_titel(),item_value)))
    """


    def __init__(self,file_path,form): #传入 path & 表单name
        self.load_file=load_workbook(file_path) #加载文件
        self.chooes_form=self.load_file[form]   #选择表单

    def  get_titel(self):
        titel=[]                                      # 创建titel
        for  items in list(self.chooes_form.rows)[0]: # for循环获取  list（表单第一行数据）[0]
            titel.append(items.value)
        return titel                                  #把循环的值返回

    def all_data(self):
        all_data=[]
        for itemsa in list(self.chooes_form.rows)[1:]:
            item_value=[]
            for  values in itemsa:
                item_value.append(values.value)
            all_data.append(dict(zip(self.get_titel(),item_value)))
        return all_data


class Loggingg(logging.Logger):
    '''
    1.创建Loggingg(logging.Logger)类 继承logger
        2.属性初始化 （name，lever，file=None，format，）
        3.supre()重新父类属性
    # 创建收集器
    #创建控制台渠道
    #设置格式
    #渠道绑定格式
    #收集器绑定渠道
    '''

    def  __init__(self,
                  name='shiyong',
                  lever='DEBUG',
                  file=None,
                  format="%(filename)s:%(lineno)d:%(name)s:%(levelname)s:%(message)s"):
        super().__init__(name,lever)
        # super(Loggingg, self).__init__(name,lever)

        # 创建收集器
        logger = logging.getLogger(name)

        logger.setLevel(logging.WARNING)
        #创建控制台渠道
        concel_handel=logging.StreamHandler()

        #设置格式
        format=logging.Formatter(format)
        #渠道绑定格式
        concel_handel.setFormatter(format)
        #收集器绑定渠道
        self.addHandler(concel_handel)

        if file:
            # 创建文件渠道
            file_handel = logging.FileHandler(file,mode='w',encoding='utf-8')
            file_handel.setLevel(lever)
            file_handel.setFormatter(format)
            self.addHandler(file_handel)
new_login=Loggingg(name='log',file=config.log_dir+r'\new_log.txt')


class Database:
    '''
    1.创建Database类
        2.数据初始化 连结属性 pymysql.connect(（host，port，user,password,database,charset）
        3.获取游标 cursor 返回字典数据：self.connect.cursor(pymysql.cursors.DictCursor)
        4.定义 方法
         传入sq，执行sq
         返回 游标.fachall、one 数据
         返回影响行数
         刷新数据 connect.commit（）

    '''

    #创建连结
    # def __init__(self,host,port,user,password,database):
    #     self.connect=pymysql.connect(
    #         host=host,
    #         port=port,
    #         user=user,
    #         password=password,
    #         database=database,
    #         charset='utf8',
    #     )
    def __init__(self):
        self.connect=pymysql.connect(
            host='rm-uf60nj0t33i3601vx3o.mysql.rds.aliyuncs.com',
            port=3306,
            user='root',
            password='shi1557225637_',
            database='table_one',
            charset='utf8',

        )
        #获取游标
        self.cur=self.connect.cursor(pymysql.cursors.DictCursor)

    def get_fetchall(self,sq):
        self.cur.execute(sq)
        return self.cur.fetchall()

    def get_fetchone(self,sq):
        self.cur.execute(sq)
        return self.cur.fetchone()

    def count_line(self,sq):
        count=self.cur.execute(sq)
        return count

    def commit(self):
        self.connect.commit()




class Config_file:
    '''
    1.from configparser import ConfigParser
    2.创建 Config_file类
        属性初始化（file_path）
        创建ConfigParser()对象
        用对象 read方法读取文件
    3.创建方法
        get_str
        get_bool
        get_float
        get_int

    '''
    def __init__(self,file_path):
        self.new_config_parser=ConfigParser()
        self.red_file=self.new_config_parser.read(file_path,'utf-8')

    def get_str(self,section,option):
        return self.new_config_parser.get(section,option)

    def get_bool(self,section,option):
        return self.new_config_parser.getboolean(section,option)

    def get_float(self,section,option):
        return self.new_config_parser.getfloat(section,option)

    def get_int(self,section,option):
        return self.new_config_parser.getint(section,option)



# def False_data():
#     db=Database()
#     # sq="INSERT INTO users  (NAME,PHONE,EMAIL) VALUES('石现涛',{},'1557225637@qq.com')".format(phone)
#     sq1="select  * FROM  users  where  PHONE=1628554195"
#     count=db.count_line(sq1)
#     # db.commit()
#     print(count)
# False_data()



# sq="INSERT INTO users  (NAME,PHONE,EMAIL) VALUES('石现涛',11111111111,'1557225637@qq.com')"
# sq1="DELETE FROM users  WHERE NAME='石现涛'"
# db=Database()
# con=db.count_line(sq1)
# db.commit()
# print(con)
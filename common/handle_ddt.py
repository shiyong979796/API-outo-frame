#:@ TIME 2021/10/23   22:03
#:@FILE  handle_ddt.py
#:@EMAIL  1557225637@QQ.COM
from openpyxl import load_workbook


import os
import config

class  Global_var:
    pass


class  Ddt_data:
    """
    1.导入
        DDT   data driver
        form  openpyxl  import  load_workbook
    2.创建类
        初始化数据  init  file_path  &  form

            加载文件 load_workbook(file_path)
            选择表单 self.load_file[form]
    3.创建方法
        获取表单第一行数据
             创建titel列表
             for循环获取  list（表单第一行数据）[0]
             把循环到的值 添加到列表
             返回出去

        获取每行的数据 以key value的形式返回
            创建 总数据列表
            for 循环获取 list（表单）[1:]
            创建行列表
            把循环到的行值添加到行列表内
            用title 做key 循环到的value 做值 all_data.append(dict(zip(self.get_titel(),item_value)))
    """
    def __init__(self,file_path,form):
        self.load_file=load_workbook(file_path)
        self.chooes_form=self.load_file[form]

    def  get_titel(self):
        titel=[]                                      #
        for  items in list(self.chooes_form.rows)[0]:
            titel.append(items.value)
        return titel

    def all_data(self):
        all_data=[]
        for itemsa in list(self.chooes_form.rows)[1:]:
            item_value=[]
            for  values in itemsa:
                item_value.append(values.value)
            all_data.append(dict(zip(self.get_titel(),item_value)))
        return all_data


    def close_excel(self):
        self.load_file.close()


new_register_data=Ddt_data(os.path.join(config.data_dir,'register.xlsx'),'register_form')#创建ddt对象



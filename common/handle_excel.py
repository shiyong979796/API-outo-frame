#:@ TIME 2021/10/23   22:03
#:@FILE  handle_excel.py
#:@EMAIL  1557225637@QQ.COM
from openpyxl import load_workbook
from common.handle_data import Global_var

class  Excel_data:
    """
    1.导入
        DDT   data driver
        form  openpyxl  import  load_workbook
    2.创建类
        初始化数据  init  file_path  &  form

            加载文件 load_workbook(file_path)
            选择表单 self.load_file[form]
    3.创建方法
        获取表单第一行数据
             创建titel列表
             for循环获取  list（表单第一行数据）[0]
             把循环到的值 添加到列表
             返回出去

        获取每行的数据 以key value的形式返回
            创建 总数据列表
            for 循环获取 list（表单）[1:]
            创建行列表
            把循环到的行值添加到行列表内
            用title 做key 循环到的value 做值 all_data.append(dict(zip(self.get_titel(),item_value)))
    """
    #初始化数据  init  file_path  &  form
    def __init__(self,file_path,form):
        self.load_file=load_workbook(file_path)
        self.chooes_form=self.load_file[form]

    def  get_titel(self):
        titel=[]                                      #
        for  items in list(self.chooes_form.rows)[0]:
            titel.append(items.value)
        return titel

    def all_data(self):
        all_data=[]
        for itemsa in list(self.chooes_form.rows)[1:]:  # 遍历数据行
            item_value=[]
            for  values in itemsa: # 获取每一行的值
                item_value.append(values.value)
            res=dict(zip(self.get_titel(),item_value)) # title和每一行数据，打包成字典
            all_data.append(res)
        return all_data


    def close_excel(self):
        self.load_file.close()

# if __name__ == '__main__':



